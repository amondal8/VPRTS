import numpy as np
import openpyxl as op

filepath= "/Thesis/Files/Mapping.xlsx"
workbook = op.load_workbook(filepath)
worksheetReq = workbook["Req_Mapping"]
worksheetR1 = workbook["R1"]
worksheetReq_rev = workbook["Req_Mapping_Rev"]
r1_prefix = "US1"
totaltc_count=20
usp_threshold=0


def createdictfromexcel(filepath, worksheetName, keycol, valuecol):
    workbook = op.load_workbook(filepath)
    worksheet = workbook[worksheetName]
    maxr=worksheet.max_row
    keyval_dict= {}

    for row_ind, row in enumerate(worksheet.iter_rows(min_row=1, max_row=maxr, min_col=keycol, max_col=keycol), start=1):
        for cell in row:
            valcell=worksheet.cell(row=row_ind,column=valuecol)
            val=valcell.value
            # if value is None:
            #     value=None
            keyval_dict[cell.value]=val
    # print(keyval_dict)
    return keyval_dict


def creation_userstorysubsets(usp_threshold):

    maxr=worksheetR1.max_row
    key = ""
    value = ""
    us_dict={}
    us_list=[]
    row_count=2
    for row in worksheetR1.iter_rows(min_row=2, max_row=maxr, min_col=1, max_col=1):
        for cell in row:
            key = cell.value
            val_cell = worksheetR1.cell(row=row_count, column=4)
            if val_cell.value is None:
                value = 0
            else:
                value = val_cell.value
        us_dict[key] = value
        row_count += 1
    # print(us_dict)
    for key in us_dict:
        if us_dict.get(key)>usp_threshold:
            us_list.append(str(key))
    print(f"User stories that are selected based on the threashold value of {usp_threshold}: {us_list}")
    return us_list


def creation_testcasesubset(us_list):

    maxr = worksheetReq.max_row
    final_row = 0
    row_count = 0
    my_tcSet = set()
    for x in us_list:
        final_row = 0
        row_count = 0
        for row in worksheetReq.iter_rows(min_row=1, max_row=maxr, min_col=1, max_col=1):
            row_count += 1
            for cell in row:
                if(cell.value == r1_prefix+str(x)):
                    final_row = row_count
                    break
            if(final_row != 0):
                break
        col_count=worksheetReq.max_column
        for row in worksheetReq.iter_rows(min_row=final_row, max_row=final_row, min_col=2, max_col=col_count):
            for cell in row:
                if(cell.value is None):
                    break
                else:
                    my_tcSet.add(cell.value)
    print(f"Printing the selected set of test cases based on the user story selection: {sorted(my_tcSet)}")
    print(f"Number of test cases selected = {len(my_tcSet)} out of a total of {totaltc_count} test cases")
    print(f"Selection rate: {(len(my_tcSet)/totaltc_count)*100}")
    return sorted(my_tcSet)


def creationof_prioritydict(tc_set):
    maxr = worksheetReq_rev.max_row
    maxc = worksheetReq_rev.max_column
    set_row=0
    row_count=0
    usp_val=0
    tc_dict={}
    for i in tc_set:
        for row_ind, row1 in enumerate(worksheetReq_rev.iter_rows(min_row=1, max_row=maxr, min_col=1, max_col=1),start=1):
            for cell1 in row1:
                if(cell1.value==i):
                    usp_val = 0
                    for row2 in worksheetReq_rev.iter_rows(min_row=row_ind, max_row=row_ind, min_col=2, max_col=maxc):
                        for cell2 in row2:
                            if (cell2.value is None):
                                break
                            else:
                                val=cell2.value[3:]
                                # print(val)
                                temp_rwcount=1
                                flag=False
                                for row3 in worksheetR1.iter_rows(min_row=2, max_row=worksheetR1.max_row, min_col=1, max_col=1):
                                    temp_rwcount+=1
                                    for cell3 in row3:
                                        if str(cell3.value) == val:
                                            # print("i am inside")
                                            usp_cell=worksheetR1.cell(row=temp_rwcount, column=3)
                                            # print(type(usp_cell.value))
                                            usp_val=usp_val+usp_cell.value
                                            flag=True
                                            break
                                    if flag==True:
                                        break
                    tc_dict[i]=usp_val

    sorted_tcdict = dict(sorted(tc_dict.items(), key=lambda x: x[1], reverse=True))
    print(f"Printing the selected tc set with usp: {tc_dict}")
    print(f"Printing the selected set in decreasing order of usp: {sorted_tcdict}")
    return sorted_tcdict

def creatingtcset_fixedexecutiontime(sortedtc_dict, total_exectime, tc_exectime):
    tcset_fixedexectime=set()
    total_time = total_exectime
    for key in sortedtc_dict:
        if total_exectime >= tc_exectime:
            tcset_fixedexectime.add(key)
            total_exectime -= tc_exectime
        else:
            break
    print(f"Subset selected based on the total execution time of {total_time} minutes where each test case is having a fixed execution time of {tc_exectime} minutes: {tcset_fixedexectime}")
    print(f"Total test cases selected with fixed execution time of test cases= {len(tcset_fixedexectime)} out of {len(sortedtc_dict)}")
    return tcset_fixedexectime


def creatingtcset_varyingecutiontime(sortedtc_dict, total_exectime, keyval_dict):
    # There must be a mapping sheet to map the tcs with the execution time
    tcset_varyingexectime=set()
    tc_list=[]
    usp_list=[]
    tcexectime_list=[]
    finaltc_list=[]
    capacity = total_exectime
    for key in sortedtc_dict:
        tc_list.append(key)
        usp_list.append(sortedtc_dict.get(key))
        tcexectime_list.append(keyval_dict.get(key))
    n = len(tcexectime_list)
    dp = [[0] * (capacity + 1) for _ in range(n + 1)]

    for i in range(1, n + 1):
        for w in range(1, capacity + 1):
            if tcexectime_list[i - 1] <= w:
                dp[i][w] = max(usp_list[i - 1] + dp[i - 1][w - tcexectime_list[i - 1]], dp[i - 1][w])
            else:
                dp[i][w] = dp[i - 1][w]

    selected_items = []
    w = capacity
    for i in range(n, 0, -1):
        if dp[i][w] != dp[i - 1][w]:
            selected_items.append(i - 1)
            w -= tcexectime_list[i - 1]

    for i in selected_items:
        finaltc_list.append(tc_list[i])

    print(f"list of tcs with varying exectime: {sorted(finaltc_list)} and capacity: {dp[n][capacity]} story points out of {sum(sortedtc_dict.values())} story points")
    print(f"Total tcs selected for varying execution time with an execution window of {total_exectime} mins: {len(finaltc_list)}")
    return dp[n][capacity], sorted(finaltc_list)

    # print(f"Subset selected based on the total execution time of {total_time} minutes where each test case is having a fixed execution time of {tc_exectime} minutes: {tcset_fixedexectime}")
    # print(f"Total test cases selected = {len(tcset_fixedexectime)} out of {len(sortedtc_dict)}")

# list=[2,7,8,10]


# list=creation_userstorysubsets()
# tc_set=creation_testcasesubset(list)
# sorted_tcdict = creationof_prioritydict(tc_set)
# creatingtcset_fixedexecutiontime(sorted_tcdict,89,15)
#
# tcexec_dict=createdictfromexcel(filepath, "TC_Executiontime", 1, 2)
# creatingtcset_varyingecutiontime(sorted_tcdict, 89, tcexec_dict)

# ************************creation of priority dict*****************************
# mdict={'us1': 128.45, 'us2': 45.68, 'us3':167.78}
# print(mdict)
# sorted_dict=dict(sorted(mdict.items(),key=lambda x:x[1], reverse=True))
# print(sorted_dict)
# for key in sorted_dict:
#     print(sorted_dict.get(key))