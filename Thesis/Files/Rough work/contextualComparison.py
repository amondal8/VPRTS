import openpyxl as op
import spacy

model1 = "en_core_web_lg"
filepath= "../Excel Implementation/Mapping.xlsx"
try:
    nlp = spacy.load(model1)
except OSError:
    from spacy.cli import download
    download(model1)
    nlp = spacy.load(model1)


def textsimilarity(text1,text2):
    doc1 = nlp(text1)
    doc2 = nlp(text2)
    return doc1.similarity(doc2)


def writeto_excel(cell, value):
    cell.value = value
    workbook.save(filepath)

def buildmatrixsheet(workbook):
    worksheetmat=workbook["Matrix"]
    worksheet1 = workbook["R1"]
    worksheet2 = workbook["R2"]
    rowcount=3
    colcount=2
    maxr1 = worksheet1.max_row
    maxr2 = worksheet2.max_row

    for row2 in worksheet2.iter_rows(min_row=2, max_row=maxr2, min_col=1, max_col=1):
        for cell2 in row2:
            r2 = cell2.value
            cell = worksheetmat.cell(row=rowcount, column=colcount)
            cell.value = "R2"+str(r2)
        rowcount+=1
    workbook.save(filepath)
    rowcount = 2
    colcount = 3
    for row1 in worksheet1.iter_rows(min_row=2, max_row=maxr1, min_col=1, max_col=1):
        for cell1 in row1:
            r1 = cell1.value
            cell = worksheetmat.cell(row=rowcount, column=colcount)
            cell.value = "R1"+str(r1)
        colcount+=1
    workbook.save(filepath)

def writeuspointsto_matsheet(workbook):
    worksheetmat = workbook["Matrix"]
    worksheet1 = workbook["R1"]
    worksheet2 = workbook["R2"]
    rowcount = 3
    colcount = 1
    maxr1 = worksheet1.max_row
    maxr2 = worksheet2.max_row

    for row2 in worksheet2.iter_rows(min_row=2, max_row=maxr2, min_col=3, max_col=3):
        for cell2 in row2:
            r2 = cell2.value
            cell = worksheetmat.cell(row=rowcount, column=colcount)
            cell.value = r2
        rowcount += 1
    workbook.save(filepath)
    rowcount = 1
    colcount = 3
    for row1 in worksheet1.iter_rows(min_row=2, max_row=maxr1, min_col=3, max_col=3):
        for cell1 in row1:
            r1 = cell1.value
            cell = worksheetmat.cell(row=rowcount, column=colcount)
            cell.value = r1
        colcount += 1
    workbook.save(filepath)


def writematrix(workbook,mat):
    worksheetmat=workbook["Matrix"]
    totrows=len(mat)
    totcols=len(mat[0])
    sheetrowcount = 2
    sheetcolcount = 3

    for row in range(totrows):
        # print(f"row: {sheetrowcount}")
        sheetrowcount += 1
        sheetcolcount = 3
        for col in range(totcols):
            cell = worksheetmat.cell(row=sheetrowcount, column=sheetcolcount)
            cell.value = mat[row][col]
            # print(f"col: {sheetcolcount}")
            sheetcolcount += 1
    workbook.save(filepath)


workbook = op.load_workbook(filepath)
worksheet1 = workbook["R1"]
worksheet2 = workbook["R2"]
worksheetMap = workbook["Mapping"]
maxr1 = worksheet1.max_row
maxr2 = worksheet2.max_row
print(f"r1: {maxr1} r2: {maxr2}")
mat=[[0 for i in range(maxr1-1)] for j in range(maxr2-1)]
rwCount1 = 1
rwCount2 = 1
similarity_threshold = 0.85
mappingValue = ""
for row2 in worksheet2.iter_rows(min_row=2, max_row=maxr2, min_col=2, max_col=2):
    for cell2 in row2:
        r2 = cell2.value
        rwCount2 += 1
        if rwCount2 <= 2:
            pass
        else:
            rwCount1 = 1
        mappingValue = ""
        for row1 in worksheet1.iter_rows(min_row=2, max_row=maxr1, min_col=2, max_col=2):
            for cell1 in row1:
                r1 = cell1.value
                rwCount1 += 1
                print(f"R2: {rwCount2} and R1: {rwCount1}")
                similarity_score = textsimilarity(r1, r2)
                mat[rwCount2-2][rwCount1-2] = similarity_score
                if similarity_score > similarity_threshold:
                    # print(f"score: {similarity_score} ")      ********
                          # f"r1:{r1} and r2:{r2}")
                    userStory_PointR2 = int(worksheet2.cell(row=rwCount2, column=3).value)
                    if worksheet1.cell(row=rwCount1, column=3).value is None:
                        userStory_PointR1 = 0
                    else:
                        usp = worksheet1.cell(row=rwCount1, column=3).value
                        userStory_PointR1 = int(usp)

                    userStory_PointR1 = userStory_PointR1+userStory_PointR2
                    cell = worksheet1.cell(row=rwCount1, column=3)
                    writeto_excel(cell, userStory_PointR1)

                    mapVal2 = worksheet2.cell(row=rwCount2, column=1).value
                    cell = worksheetMap.cell(row=rwCount2, column=1)
                    writeto_excel(cell, mapVal2)

                    mapVal1 = worksheet1.cell(row=rwCount1, column=1).value
                    if len(mappingValue) == 0:
                        mappingValue += str(mapVal1)
                    else:
                        mappingValue = mappingValue+',' + str(mapVal1)
                    cell = worksheetMap.cell(row=rwCount2, column=2)
                    writeto_excel(cell, mappingValue)

                    # print(f"Updated userStory_PointR1: {userStory_PointR1}")

                else:
                    # print(f"Similarity value below the threshold {similarity_threshold}")     **********
                    mapVal2 = worksheet2.cell(row=rwCount2, column=1).value
                    cell = worksheetMap.cell(row=rwCount2, column=1)
                    writeto_excel(cell, mapVal2)

                    # cell = worksheetMap.cell(row=rwCount2, column=2)
                    # writeto_excel(cell, "No Match")

print(mat)
writematrix(workbook,mat)
buildmatrixsheet(workbook)
writeuspointsto_matsheet(workbook)






















# textn1="As a developer, I'd like to move k8s SPI to it's own repo."
# textn2="As a developer, I'd like to upgrade Spring XD's ambari plugin to 1.3 release."
#
# print(textsimilarity(textn1,textn2))