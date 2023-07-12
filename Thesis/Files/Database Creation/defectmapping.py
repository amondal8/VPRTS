import numpy as np
import openpyxl as op
import random

filepath= "C:/Users/amondal8/PycharmProjects/pythonProject3/Thesis/Files/Database Creation/Mapping.xlsx"
workbook = op.load_workbook(filepath)
worksheetDef = workbook["Defect_Mapping"]
totaldefect_count=10



def generate_defectmap_adjmat_withlimits(tc_count,defect_count,connection_prob,lower_limitones,upper_limitones):

    n_rows = tc_count  # Number of rows
    n_cols = defect_count  # Number of columns


    defect_matrix = np.zeros((n_rows, n_cols), dtype=int)  # Initialize an all-zero adjacency matrix

    for i in range(n_rows):
        max_ones_per_row = random.randint(lower_limitones, upper_limitones)    # Maximum number of ones per row
        ones_count = np.sum(defect_matrix[i])  # Count the number of ones in the current row
        if ones_count < max_ones_per_row and np.random.rand() < connection_prob:
            remaining_ones = max_ones_per_row - ones_count
            available_indices = np.where(defect_matrix[i] == 0)[0]  # Get indices of zeros in the row
            num_indices = min(remaining_ones, len(available_indices))
            random_indices = np.random.choice(available_indices, size=num_indices, replace=False)
                                              # p=[connection_probability] * num_indices)
            defect_matrix[i, random_indices] = 1

    return defect_matrix

def writematto_defectmappingexcel(adj_mat):

    workbook = op.load_workbook(filepath)
    worksheet = workbook["Defect_Mapping"]
    us_count = 1
    tc_count = 1
    col_counter = 2

    for x in adj_mat:
        # print(x)
        cell = worksheet.cell(row=us_count,column=1)
        cell.value = "TC"+str(us_count)
        for y in x:
            # print(y)
            if (y == 1):
                cell = worksheet.cell(row=us_count,column=col_counter)
                cell.value = "D"+str(tc_count)
                col_counter+=1
            tc_count += 1

        col_counter=2
        tc_count=1
        us_count += 1
        workbook.save(filepath)

    print("End of printing defect mapping matrix to excel")

def creation_defectsubset(tc_list):
    # filepath = "C:/Users/amondal8/PycharmProjects/pythonProject3/Thesis/Files/Mapping.xlsx"
    workbook = op.load_workbook(filepath)
    worksheetDef = workbook["Defect_Mapping"]
    maxr = worksheetDef.max_row
    final_row = 0
    row_count = 0
    my_defectSet=set()
    for x in tc_list:
        final_row = 0
        row_count = 0
        for row in worksheetDef.iter_rows(min_row=1, max_row=maxr, min_col=1, max_col=1):
            row_count += 1
            for cell in row:
                if(cell.value==str(x)):
                    final_row = row_count
                    break
            if(final_row != 0):
                break
        col_count=worksheetDef.max_column
        for row in worksheetDef.iter_rows(min_row=final_row, max_row=final_row, min_col=2, max_col=col_count):
            for cell in row:
                if(cell.value is None):
                    break
                else:
                    my_defectSet.add(cell.value)
    print(f"Printing the identified set of defects based on the selected subset of test cases: {sorted(my_defectSet)}")
    print(f"Number of defects identified = {len(my_defectSet)} out of a total of {totaldefect_count} defects")
    return sorted(my_defectSet)

def creation_defectdict():
    # filepath = "C:/Users/amondal8/PycharmProjects/pythonProject3/Thesis/Files/Mapping.xlsx"
    workbook = op.load_workbook(filepath)
    worksheetDef = workbook["Defect_Mapping"]
    res_dict = {}
    temp_set=set()
    maxr=worksheetDef.max_row
    maxc=worksheetDef.max_column
    for x_count, x in enumerate(worksheetDef.iter_rows(min_row=1, max_row=maxr, min_col=2, max_col=maxc),start=1):
        key_cell=worksheetDef.cell(row=x_count,column=1)
        key=key_cell.value
        for y in x:
            if y.value is None:
                break
            else:
                temp_set.add(y.value)
                # print(temp_set)
        res_dict[key] = set(temp_set)
        temp_set.clear()
    print(res_dict)
    return res_dict

def defect_issubset(defect_dict, defect_subset):        # This will not make any sence as this will also mean that we
    # are preserving the values for the tcs which do not have any defects which is wrong
    additional_testcases=set()
    for key in defect_dict:
        if set(defect_dict.get(key)).issubset(set(defect_subset)):
            additional_testcases.add(key)
    print(f"The additional set of tc is: {additional_testcases}")
    return sorted(additional_testcases)




# defectmap_adjmat=generate_defectmap_adjmat_withlimits(50,10,.5,0,3)
# # print(defectmap_adjmat)
# writematto_defectmappingexcel(defectmap_adjmat)
# # tc_list=["TC1","TC2","TC3"]
# myset=creation_defectsubset(tc_list)
# # trial_set=["D9","D2","D4"]
# # print(set(trial_set).issubset(set(myset)))
# defect_dict=creation_defectdict()
# defect_issubset(defect_dict,myset)

