import mysql.connector
import openpyxl as op
import connectionclass as cc


mydb = mysql.connector.connect(
  host="127.0.0.1",
  user="root",
  password="Aa*231491*dD",
  database="new1"
)

#Add your credentials/path
filepath= "/Thesis/Files/Extra/Database Creation/Mapping.xlsx"
workbook = op.load_workbook(filepath)

def gettrack_id():
  mycursor = mydb.cursor()
  mycursor.execute("Select id from idtracker ORDER BY timestamp DESC LIMIT 1")
  result = mycursor.fetchone()
  if result is None:
    return result
  else:
    return result[0]


def stringcreation_columns(my_list):
  my_str = ""
  for ind, i in enumerate(my_list):
    if ind == 0:
      my_str = my_str + "(" + str(i)
    elif ind == len(my_list) - 1:
      my_str = my_str + "," + str(i) + ")"
    else:
      my_str = my_str + "," + str(i)
  return my_str


def stringcreation_values(my_list, datatype_list):
  my_str = ""
  for ind, i in enumerate(my_list):
    if ind == 0:
      if datatype_list[ind] == "int":
        my_str = my_str + "(" + str(i)
      else:
        my_str = my_str + "(\"" + str(i) + "\""
    elif ind == len(my_list) - 1:
      if datatype_list[ind] == "int":
        my_str = my_str + "," + str(i) + ")"
      else:
        my_str = my_str + ",\"" + str(i) + "\")"
    else:
      if datatype_list[ind] == "int":
        my_str = my_str + "," + str(i)
      else:
        my_str = my_str + ",\"" + str(i) + "\""
  return my_str


def insertquery_creation(tablename, collist, vallist, datatypelist):
  cols = stringcreation_columns(collist)
  vals = stringcreation_values(vallist, datatypelist)
  query = f"""INSERT INTO {tablename} {cols} VALUES {vals}"""
  return query


def running_insertquery(query):
  mydb = mysql.connector.connect(
    host="127.0.0.1",
    user="root",
    password="Aa*231491*dD",
    database="new1"
  )
  mycursor = mydb.cursor()
  if mydb.is_closed():
    mydb.start_transaction()
  # mycursor = cc.connection()
  mycursor.execute(query)
  print(f"Query {query} executed successfully")
  mydb.commit()
  # cc.closeconnection()
  mycursor.close()
  mydb.close()


def createdictfromexcel(filepath, worksheetName, keycol, valuecol):
  workbook = op.load_workbook(filepath)
  worksheet = workbook[worksheetName]
  maxr = worksheet.max_row
  keyval_dict = {}

  for row_ind, row in enumerate(worksheet.iter_rows(min_row=1, max_row=maxr, min_col=keycol, max_col=keycol),
                                start=1):
    for cell in row:
      valcell = worksheet.cell(row=row_ind, column=valuecol)
      val = valcell.value
      # if value is None:
      #     value=None
      keyval_dict[cell.value] = val
  # print(keyval_dict)
  return keyval_dict


def fetch_ones_in_row(adjacency_matrix, row_index):
  row = adjacency_matrix[row_index]
  ones_indices = [index for index, value in enumerate(row) if value == 1]
  return ones_indices


def create_searchquery(us_list, colname, tablename, searchcolname, track_id, prefix):
  us_str = ""
  for ind, i in enumerate(us_list):
    if ind == len(us_list)-1:
      us_str = us_str + "\"" + prefix + str(i) +"\""
    else:
      us_str = us_str+"\"" + prefix + str(i) + "\","
  query = f"""SELECT DISTINCT {colname} from {tablename} where {searchcolname} in ({us_str}) and track_id = {track_id}"""
  print(query)
  return query

def running_searchqury(query):
  mycursor = mydb.cursor()
  mycursor.execute(query)
  result = mycursor.fetchall()
  # print(result)
  return result

def create_fetchquery(tc_id, track_id):
  query = f"""Select sum(us_points) from user_story where us_id in(SELECT DISTINCT us_id from us_tc_map where tc_id in ("{tc_id}") and track_id = {track_id}) and track_id = {track_id}"""
  return query


def creating_prioritydict_tclist(tc_list, track_id):
  tc_dict = {}
  for ind, i in enumerate(tc_list):
    query = create_fetchquery(i, track_id)
    # print(query)
    result = running_searchqury(query)
    tc_dict[i] = int(result[0][0])
  sorted_tcdict = dict(sorted(tc_dict.items(), key=lambda x: x[1], reverse=True))
  print(f"Printing the selected tc set with usp: {tc_dict}")
  print(f"Printing the selected set in decreasing order of usp: {sorted_tcdict}")
  return sorted_tcdict


# creating_prioritydict_tclist(["TC1", "TC2"], "1")

# us_lst = [1,2,3,4]
# create_tcsearchquery(us_lst, "tc_id", "us_tc_map", "us_id")