import numpy as np
import openpyxl as op
import random

#Add your credentials/path
filepath = "/Thesis/Files/Extra/Database Creation/Mapping.xlsx"
total_uscount=10
total_tccount=20
connection_prob=1   # Using a value less than 1 will decrease the 1s even more, so use it wisely
limiting_ones=8

def generate_random_adjmat_withlimits(n,m,connection_prob,limit_ones):

    n_rows = n  # Number of rows
    n_cols = m  # Number of columns
    # max_ones_per_row = limit_ones  # Maximum number of ones per row
    # connection_prob = Probability of connection (0.0 to 1.0)

    adj_matrix = np.zeros((n_rows, n_cols), dtype=int)  # Initialize an all-zero adjacency matrix

    for i in range(n_rows):
        max_ones_per_row = random.randint(5, limit_ones)    # Maximum number of ones per row
        ones_count = np.sum(adj_matrix[i])  # Count the number of ones in the current row
        if ones_count < max_ones_per_row and np.random.rand() < connection_prob:
            remaining_ones = max_ones_per_row - ones_count
            available_indices = np.where(adj_matrix[i] == 0)[0]  # Get indices of zeros in the row
            num_indices = min(remaining_ones, len(available_indices))
            random_indices = np.random.choice(available_indices, size=num_indices, replace=False)
                                              # p=[connection_probability] * num_indices)
            adj_matrix[i, random_indices] = 1

    return adj_matrix


def writematto_reqmappingexcel(adj_mat):

    workbook = op.load_workbook(filepath)
    worksheet = workbook["Req_Mapping"]
    us_count = 1
    tc_count = 1
    col_counter = 2

    for x in adj_mat:
        # print(x)
        cell = worksheet.cell(row=us_count,column=1)
        cell.value = "US1"+str(us_count)
        for y in x:
            # print(y)
            if (y == 1):
                cell = worksheet.cell(row=us_count,column=col_counter)
                cell.value = "TC"+str(tc_count)
                col_counter+=1
            tc_count += 1

        col_counter=2
        tc_count=1
        us_count += 1
        workbook.save(filepath)
        workbook.close()

    print("End of printing requirement mapping matrix to excel")

def writeto_revReqMat_excel(adj_matrix):
    workbook = op.load_workbook(filepath)
    worksheet = workbook["Req_Mapping_Rev"]
    row_count=1
    col_count=2
    for i in range(len(adj_matrix[0])):
        # available_indices = np.where(adj_matrix[i] == 1)[0]
        indices = np.where(adj_matrix[:, i] == 1)[0]
        # print(indices)
        cell = worksheet.cell(row=row_count, column=1)
        cell.value = "TC" + str(row_count)
        for x in indices:
            cell = worksheet.cell(row=row_count, column=col_count)
            cell.value = "US1" + str(x+1)
            col_count+=1
        col_count=2
        row_count+=1
    workbook.save(filepath)
    workbook.close()
    print("End of printing reverse requirement mapping to excel")

# adj_matrix = generate_random_adjmat_withlimits(total_uscount,total_tccount,connection_prob,limiting_ones)
# print(adj_matrix)
# writematto_reqmappingexcel(adj_matrix)
# writeto_revReqMat_excel(adj_matrix)


