import numpy as np
import openpyxl as op
import random

filepath = "/Thesis/Files/Extra/Database Creation/Mapping.xlsx"
total_uscount=10
total_tccount=20
connection_prob=1   # Using a value less than 1 will decrease the 1s even more, so use it wisely
limiting_ones=8

def generate_random_adjmat_withlimits(n,m,connection_prob,limit_ones):

    n_rows = n  # Number of rows
    n_cols = m  # Number of columns
    # max_ones_per_row = limit_ones  # Maximum number of ones per row
    # connection_prob = Probability of connection (0.0 to 1.0)

    adj_matrix = np.zeros((n_rows, n_cols), dtype=int)  # Initialize an all-zero adjacency matrix

    for i in range(n_rows):
        max_ones_per_row = random.randint(5, limit_ones)    # Maximum number of ones per row
        ones_count = np.sum(adj_matrix[i])  # Count the number of ones in the current row
        if ones_count < max_ones_per_row and np.random.rand() < connection_prob:
            remaining_ones = max_ones_per_row - ones_count
            available_indices = np.where(adj_matrix[i] == 0)[0]  # Get indices of zeros in the row
            num_indices = min(remaining_ones, len(available_indices))
            random_indices = np.random.choice(available_indices, size=num_indices, replace=False)
                                              # p=[connection_probability] * num_indices)
            adj_matrix[i, random_indices] = 1

    return adj_matrix

def generate_adjmat_onetomany_withlimits(num_rows, num_columns, limiting_ones, connection_probability):

    limiting_ones = int(num_columns/num_rows) + 1
    print(f"limiting_ones: {limiting_ones}")
    # Initialize an empty adjacency matrix
    adjacency_matrix = [[0 for _ in range(num_columns)] for _ in range(num_rows)]

    # Keep track of "many" entities that have already been connected to a "one" entity
    used_many_entities = set()

    # Fill in the matrix based on the connection probability and max_ones_per_row
    for one_index in range(num_rows):
        max_ones_per_row = random.randint(1, limiting_ones)  # Maximum number of ones per row
        ones_count = 0
        for many_index in range(num_columns):
            if many_index in used_many_entities:
                continue
            if ones_count >= max_ones_per_row:
                break
            if random.random() <= connection_probability:
                adjacency_matrix[one_index][many_index] = 1
                used_many_entities.add(many_index)
                ones_count += 1

    return adjacency_matrix


def generate_adjmat_onetomany_withconfig(num_rows, num_columns, limiting_ones, config, connection_probability=1.0):

    # Initialize an empty adjacency matrix
    adjacency_matrix = [[0 for _ in range(num_columns)] for _ in range(num_rows)]
    row_low = 0
    row_high = 0
    if config.lower() == "top":
        row_low = 0
        row_high = int(num_rows/4)
    elif config.lower() == "bottom":
        row_low = int((3*num_rows)/4)
        row_high = num_rows-1
    elif config.lower() == "center":
        row_low = int((num_rows) / 2)
        row_high = int((3 * num_rows) / 4)
    upper_lim = int(num_columns/6)
    print(f"row_low: {row_low}")
    print(f"row_high: {row_high}")
    print(f"upper_lim: {upper_lim}")

    # Keep track of "many" entities that have already been connected to a "one" entity
    used_many_entities = set()

    # Fill in the matrix based on the connection probability and max_ones_per_row
    for one_index in range(num_rows):
        if (one_index >= row_low and one_index < row_high) and (config.lower() == "top" or config.lower() == "bottom" or config.lower() == "center"):
            max_ones_per_row = upper_lim
        else:
            max_ones_per_row = random.randint(1, limiting_ones)  # Maximum number of ones per row
        # print(f"max_ones_per_row: {max_ones_per_row}")
        ones_count = 0
        for many_index in range(num_columns):
            if many_index in used_many_entities:
                continue
            if ones_count >= max_ones_per_row:
                break
            if random.random() <= connection_probability:
                adjacency_matrix[one_index][many_index] = 1
                used_many_entities.add(many_index)
                ones_count += 1

    return adjacency_matrix

def writematto_reqmappingexcel(adj_mat):

    workbook = op.load_workbook(filepath)
    worksheet = workbook["Req_Mapping"]
    us_count = 1
    tc_count = 1
    col_counter = 2

    for x in adj_mat:
        # print(x)
        cell = worksheet.cell(row=us_count,column=1)
        cell.value = "US1"+str(us_count)
        for y in x:
            # print(y)
            if (y == 1):
                cell = worksheet.cell(row=us_count,column=col_counter)
                cell.value = "TC"+str(tc_count)
                col_counter+=1
            tc_count += 1

        col_counter=2
        tc_count=1
        us_count += 1
        workbook.save(filepath)
        workbook.close()

    print("End of printing requirement mapping matrix to excel")

def writeto_revReqMat_excel(adj_matrix):
    # num_rows, num_columns = adj_matrix.shape
    print(f"col count: {len(adj_matrix[0])}")
    workbook = op.load_workbook(filepath)
    worksheet = workbook["Req_Mapping_Rev"]
    row_count=1
    col_count=2
    for i in range(len(adj_matrix[0])):
        # available_indices = np.where(adj_matrix[i] == 1)[0]
        # print(f"indices: {available_indices}")
        indices = np.where(adj_matrix[:, i] == 1)[0]
        print(f"indices: {type(indices)}")
        cell = worksheet.cell(row=row_count, column=1)
        cell.value = "TC" + str(row_count)
        for x in indices:
            cell = worksheet.cell(row=row_count, column=col_count)
            cell.value = "US1" + str(x+1)
            col_count+=1
        col_count=2
        row_count+=1
    workbook.save(filepath)
    workbook.close()
    print("End of printing reverse requirement mapping to excel")


def generate_tccmmap_adjmat_withlimits(us_count,cm_count,connection_prob,lower_limitones,upper_limitones):


    n_rows = us_count  # Number of rows
    n_cols = cm_count  # Number of columns


    uscm_matrix = np.zeros((n_rows, n_cols), dtype=int)  # Initialize an all-zero adjacency matrix

    for i in range(n_rows):
        max_ones_per_row = random.randint(lower_limitones, upper_limitones)    # Maximum number of ones per row
        ones_count = np.sum(uscm_matrix[i])  # Count the number of ones in the current row
        if ones_count < max_ones_per_row and np.random.rand() < connection_prob:
            remaining_ones = max_ones_per_row - ones_count
            available_indices = np.where(uscm_matrix[i] == 0)[0]  # Get indices of zeros in the row
            num_indices = min(remaining_ones, len(available_indices))
            random_indices = np.random.choice(available_indices, size=num_indices, replace=False)
                                              # p=[connection_probability] * num_indices)
            # affected_value = round(random.uniform(0, 1), 3)
            uscm_matrix[i, random_indices] = 1

    uscm_matrixfin = np.zeros((n_rows, n_cols), dtype=np.float64)
    uscm_matrixfin = uscm_matrix.astype(np.float64)
    res_mat = replace_ones_with_random(uscm_matrixfin)
    # for i in uscm_matrixfin:
    #     print(i)
    print("Mapping matrix with affected values ahve been generated between Userstory and Code modification ")
    return uscm_matrixfin


def replace_ones_with_random(matrix):
    num_rows = len(matrix)
    num_columns = len(matrix[0])

    for i in range(num_rows):
        for j in range(num_columns):
            if matrix[i][j] == 1:
                # Generate a random number between 0 and 1
                random_number = round(random.uniform(0, 1),5)
                matrix[i][j] = random_number

    return matrix

def generate_custom_adjacency_matrix(rows, columns, prob_thres, config):
    matrix = [[0 for _ in range(columns)] for _ in range(rows)]
    if config=="lefttop":
        top_left_row, top_left_column, bottom_right_row, bottom_right_column = [0,0,rows/2-1,columns/2-1]
    elif config=="righttop":
        top_left_row, top_left_column, bottom_right_row, bottom_right_column = [0,columns/2,rows/2-1,columns-1]
    elif config=="leftbottom":
        top_left_row, top_left_column, bottom_right_row, bottom_right_column = [rows/2,0,rows-1,columns/2-1]
    elif config=="rightbottom":
        top_left_row, top_left_column, bottom_right_row, bottom_right_column = [rows/2,columns/2,rows-1,columns-1]
    elif config=="leftbound":
        top_left_row, top_left_column, bottom_right_row, bottom_right_column = [0,0,rows-1,columns/2-1]
    elif config=="rightbound":
        top_left_row, top_left_column, bottom_right_row, bottom_right_column = [0,columns/2,rows-1,columns-1]
    elif config=="center":
        top_left_row, top_left_column, bottom_right_row, bottom_right_column = [rows / 4 - 1, columns / 4 - 1,
                                                                            (3 * rows) / 4, (3 * columns) / 4]
    elif config=="topbound":
        top_left_row, top_left_column, bottom_right_row, bottom_right_column = [0,0,rows/2-1,columns-1]
    elif config == "bottombound":
        top_left_row, top_left_column, bottom_right_row, bottom_right_column = [rows/2, 0, rows- 1, columns - 1]


    for i in range(rows):
        for j in range(columns):
            if top_left_row <= i <= bottom_right_row and top_left_column <= j <= bottom_right_column:
                # Set ones in the specified area
                matrix[i][j] = 1
            else:
                # Set zeros outside the specified area
                prob = np.random.rand()
                if(prob<prob_thres):
                    matrix[i][j] = 0
                else:
                    matrix[i][j] = 1
    return matrix


def generate_custom_adjacency_matrix_uscm(rows, columns, prob_thres, config):
    matrix = [[0 for _ in range(columns)] for _ in range(rows)]
    if config=="lefttop":
        top_left_row, top_left_column, bottom_right_row, bottom_right_column = [0,0,rows/2-1,columns/2-1]
    elif config=="righttop":
        top_left_row, top_left_column, bottom_right_row, bottom_right_column = [0,columns/2,rows/2-1,columns-1]
    elif config=="leftbottom":
        top_left_row, top_left_column, bottom_right_row, bottom_right_column = [rows/2,0,rows-1,columns/2-1]
    elif config=="rightbottom":
        top_left_row, top_left_column, bottom_right_row, bottom_right_column = [rows/2,columns/2,rows-1,columns-1]
    elif config=="leftbound":
        top_left_row, top_left_column, bottom_right_row, bottom_right_column = [0,0,rows-1,columns/2-1]
    elif config=="rightbound":
        top_left_row, top_left_column, bottom_right_row, bottom_right_column = [0,columns/2,rows-1,columns-1]
    elif config=="center":
        top_left_row, top_left_column, bottom_right_row, bottom_right_column = [rows / 4 - 1, columns / 4 - 1,
                                                                            (3 * rows) / 4, (3 * columns) / 4]
    elif config == "topbound":
        top_left_row, top_left_column, bottom_right_row, bottom_right_column = [0, 0, rows / 2 - 1, columns - 1]
    elif config == "bottombound":
        top_left_row, top_left_column, bottom_right_row, bottom_right_column = [rows / 2, 0, rows - 1, columns - 1]


    for i in range(rows):
        for j in range(columns):
            val = np.random.rand()
            if top_left_row <= i <= bottom_right_row and top_left_column <= j <= bottom_right_column:
                # Set ones in the specified area
                matrix[i][j] = val
            else:
                # Set zeros outside the specified area
                prob = np.random.rand()
                if(prob>prob_thres):
                    matrix[i][j] = val
                    print(prob)
                # else:
                #     matrix[i][j] = 0
    return matrix

# adj_matrix = generate_random_adjmat_withlimits(total_uscount,total_tccount,connection_prob,limiting_ones)
# print(adj_matrix)
# writematto_reqmappingexcel(adj_matrix)
# writeto_revReqMat_excel(adj_matrix)

# mat = generate_tccmmap_adjmat_withlimits(20, 10, 1, 1, 4)
# mat1 = replace_ones_with_random(mat)
# for i in mat1:
#     print(i)

