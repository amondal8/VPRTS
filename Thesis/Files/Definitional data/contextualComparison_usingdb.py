import openpyxl as op
import spacy
import configparser
import utilities as ut
import utilities_dataset as ut_ds


filepath = "C:/Users/amondal8/PycharmProjects/pythonProject3/Thesis/Files/Database Creation/Mapping.xlsx"
config = configparser.ConfigParser()
config.read('config1.ini')
dbconnect = config['dbconnection_dataset']
data = config['data']
tablenames_config = config['tablenames']

model1 = "en_core_web_lg"
filepath = "C:/Users/amondal8/PycharmProjects/pythonProject3/Thesis/Files/Database Creation/Mapping.xlsx"
workbook = op.load_workbook(filepath)
worksheet1 = workbook["R1"]
worksheet2 = workbook["R2"]
worksheetMap = workbook["Mapping"]

fielname_towritemat = "C:/Users/amondal8/PycharmProjects/pythonProject3/Thesis/Files/Definitional data/outputMat.txt"
ustcmap_tablename = tablenames_config["ustcmap_tablename"]
uscmmap_tablename = tablenames_config["uscmmap_tablename"]
tcdefectmap_tablename = tablenames_config["tcdefectmap_tablename"]
dataset_tablename = tablenames_config["dataset_tablename"]
usdataset_tablename = tablenames_config["usdatasettable_tablename"]
tcdataset_tablename = tablenames_config["tcdatasettable_tablename"]
defectdataset_tablename = tablenames_config["defectdatasettable_tablename"]
cmdatasettable_tablename = tablenames_config["cmdatasettable_tablename"]
rel1 = data['release1']
rel2 = data['release2']
similarityval_threshold = float(data["similarityval_threshold"])

try:
    nlp = spacy.load(model1)
except OSError:
    from spacy.cli import download

    download(model1)
    nlp = spacy.load(model1)

# Configuration is the  way we want to calculate the cumulative value
# of the user story and how it is carried to the other user stories
# configuration=1        -> when the value of R2 is carried to R1
# configuration=2        -> when the value of R1 is carried to R2
# configuration=3        -> when the value of R1 + R2 is carried and placed beside R1
# configuration=4        -> when the value of R2 is carried to R1 with weight of similarity value being considered

configuration = 3

simvalue_dict = {}
usmap_dict = {}


def textsimilarity(text1, text2):
    doc1 = nlp(text1)
    doc2 = nlp(text2)
    return doc1.similarity(doc2)

def writematrix(workbook, mat, usr2_list, usr1_list):
    worksheetmat = workbook["Matrix"]
    totrows = len(mat)
    totcols = len(mat[0])
    sheetrowcount = 2
    sheetcolcount = 3
    row_head = 2
    col_head = 2

    for indi, i in enumerate(usr2_list):
        cell = worksheetmat.cell(row=indi+3, column=row_head)
        cell.value = i
    workbook.save(filepath)

    for indi, i in enumerate(usr1_list):
        cell = worksheetmat.cell(row=col_head, column=indi+3)
        cell.value = i
    workbook.save(filepath)


    for row in range(totrows):
        # print(f"row: {sheetrowcount}")
        sheetrowcount += 1
        sheetcolcount = 3
        for col in range(totcols):
            cell = worksheetmat.cell(row=sheetrowcount, column=sheetcolcount)
            cell.value = mat[row][col]
            # print(f"col: {sheetcolcount}")
            sheetcolcount += 1
    workbook.save(filepath)

def contentcomparison(ds_id, importanceval_calconfig):
    print(f"The content comparison is done of the basis of configuration: {importanceval_calconfig}")
    query_r1 = f"select us_id, us_desc, us_points, us_businessvalue from {usdataset_tablename} where ds_id = '{ds_id}' and release_id in ({rel1})"
    query_r2 = f"select us_id, us_desc, us_points, us_businessvalue from {usdataset_tablename} where ds_id = '{ds_id}' and release_id in ({rel2})"
    # query_cmval = f"select sum(affected_value) from us_cm_map where us_id = '{}' and ds_id = '{ds_id}'"
    usr1_res = ut_ds.running_searchqury(query_r1)
    usr2_res = ut_ds.running_searchqury(query_r2)
    usr1_list = ut.createlist_fromdbresult(usr1_res,0)
    usr2_list = ut.createlist_fromdbresult(usr2_res, 0)
    maxr1 = len(usr1_res)
    maxr2 = len(usr2_res)
    mat = [[0 for i in range(maxr1)] for j in range(maxr2)]

    for indi, i in enumerate(usr2_res):
        for indj, j in enumerate(usr1_res):
            simval = textsimilarity(i[1], j[1])
            # print(f"us2: {i[0]}, us1: {j[0]} have sim value: {simval}")
            if simval < 0:
                simval = 0
            mat[indi][indj] = simval

            # ******************

            # if simval >= similarityval_threshold:
            #     query_cmval = f"select sum(affected_value) from us_cm_map where us_id = '{j[0]}' and ds_id = '{ds_id}'"
            #     # print(f"query: {query_cmval}")
            #     cmval_res = ut_ds.running_searchqury(query_cmval)
            #     cmval = cmval_res[0][0]
            #     # print(f"cmval is: {cmval}")
            #     if j[0] in simvalue_dict.keys():
            #         val = simvalue_dict[j[0]]
            #         if importanceval_calconfig == "3":
            #             val += (i[2]+i[3]+(cmval*(j[2]+j[3])))
            #         elif importanceval_calconfig == "1":
            #             val += cmval*(i[2] + i[3])
            #         elif importanceval_calconfig == "4":
            #             val += simval*(j[2]+j[3] + (cmval * (i[2] + i[3])))     #cmval to be calculated for R2 not R1
            #         # print(val)
            #         # print(cmval*val)
            #         simvalue_dict[j[0]] = val
            #     else:
            #         if importanceval_calconfig == "3":
            #             simvalue_dict[j[0]] = (i[2]+i[3]+(cmval*(j[2]+j[3])))
            #         elif importanceval_calconfig == "1":
            #             simvalue_dict[j[0]] = cmval*(i[2] + i[3])
            #         elif importanceval_calconfig == "4":
            #             simvalue_dict[j[0]] = simval * (j[2] + j[3] + (cmval * (i[2] + i[3])))

            # **************************

    # print(f"User stories with cumulative value after similarity measure: {simvalue_dict}")
    ut.writetonotepad('w', "", fielname_towritemat)
    for i in mat:
        ut.writetonotepad('a', str(i)+"\n", fielname_towritemat)
    result_text = ut.read_from_txt(fielname_towritemat)
    query_updateresults = ut.updatetable_query(dataset_tablename, "similarity_value", result_text, "str", ds_id)
    ut_ds.running_insertquery(query_updateresults)
    impvalue_dict = ut_ds.importanceval_fortc(ds_id, mat, usr2_list, usr1_list, importanceval_calconfig)
    writematrix(workbook, mat, usr2_list, usr1_list)
    # simval_dict = ut_ds.simval_config(mat,usr1_list,usr2_list)



    # for i in mat:
    #     print(i)
    # print(simvalue_dict)
    sorted_impvaldict = {k: v for k, v in sorted(impvalue_dict.items(), key=lambda item: item[1], reverse = True)}
    print(f"sorted us_dict: {len(sorted_impvaldict)}{sorted_impvaldict}")
    return sorted_impvaldict

# contentcomparison("1", "1.1")


def writeto_excel(cell, value):
    cell.value = value
    workbook.save(filepath)


def buildmatrixsheet(workbook):
    worksheetmat = workbook["Matrix"]
    worksheet1 = workbook["R1"]
    worksheet2 = workbook["R2"]
    rowcount = 3
    colcount = 2
    maxr1 = worksheet1.max_row
    maxr2 = worksheet2.max_row

    for row2 in worksheet2.iter_rows(min_row=2, max_row=maxr2, min_col=1, max_col=1):
        for cell2 in row2:
            r2 = cell2.value
            cell = worksheetmat.cell(row=rowcount, column=colcount)
            cell.value = "R2" + str(r2)
        rowcount += 1
    workbook.save(filepath)
    rowcount = 2
    colcount = 3
    for row1 in worksheet1.iter_rows(min_row=2, max_row=maxr1, min_col=1, max_col=1):
        for cell1 in row1:
            r1 = cell1.value
            cell = worksheetmat.cell(row=rowcount, column=colcount)
            cell.value = "R1" + str(r1)
        colcount += 1
    workbook.save(filepath)


def writeuspointsto_matsheet(workbook):
    worksheetmat = workbook["Matrix"]
    # worksheet1 = workbook["R1"]
    # worksheet2 = workbook["R2"]
    rowcount = 3
    colcount = 1
    maxr1 = worksheet1.max_row
    maxr2 = worksheet2.max_row

    for row2 in worksheet2.iter_rows(min_row=2, max_row=maxr2, min_col=3, max_col=3):
        for cell2 in row2:
            r2 = cell2.value
            cell = worksheetmat.cell(row=rowcount, column=colcount)
            cell.value = r2
        rowcount += 1
    workbook.save(filepath)
    rowcount = 1
    colcount = 3
    for row1 in worksheet1.iter_rows(min_row=2, max_row=maxr1, min_col=3, max_col=3):
        for cell1 in row1:
            r1 = cell1.value
            cell = worksheetmat.cell(row=rowcount, column=colcount)
            cell.value = r1
        colcount += 1
    workbook.save(filepath)


def valuesCarriedToR1(rwCount1, rwCount2):  # configuration=1
    userStory_PointR2 = int(worksheet2.cell(row=rwCount2, column=3).value)
    if worksheet1.cell(row=rwCount1, column=3).value is None:
        userStory_PointR1 = 0
    else:
        usp = worksheet1.cell(row=rwCount1, column=3).value
        userStory_PointR1 = int(usp)

    userStory_PointR1 = userStory_PointR1 + userStory_PointR2
    cell = worksheet1.cell(row=rwCount1, column=3)
    writeto_excel(cell, userStory_PointR1)


def valuesCarriedToR2(rwCount1, rwCount2):  # configuration=2
    userStory_PointR1 = int(worksheet1.cell(row=rwCount1, column=3).value)
    if worksheet2.cell(row=rwCount2, column=3).value is None:
        userStory_PointR2 = 0
    else:
        usp = worksheet2.cell(row=rwCount2, column=3).value
        userStory_PointR2 = int(usp)

    userStory_PointR2 = userStory_PointR1 + userStory_PointR2
    cell = worksheet2.cell(row=rwCount2, column=3)
    writeto_excel(cell, userStory_PointR2)


def valuesR1andR2Added(rwCount1, rwCount2):  # configuration=3
    userStory_PointR1 = int(worksheet1.cell(row=rwCount1, column=3).value)
    userStory_PointR2 = int(worksheet2.cell(row=rwCount2, column=3).value)
    if worksheet1.cell(row=rwCount1, column=5).value is None:
        userStory_Pointadd = 0
    else:
        usp = worksheet1.cell(row=rwCount1, column=5).value
        userStory_Pointadd = int(usp)

    userStory_Pointadd = userStory_PointR1 + userStory_PointR2 + userStory_Pointadd
    cell = worksheet1.cell(row=rwCount1, column=5)  # writing the new value on R1 sheet: repeat of R1 is getting calculated
    writeto_excel(cell, userStory_Pointadd)


def valuesR1andR2Added2(rwCount1, rwCount2):  # configuration=3
    userStory_PointR1 = int(worksheet1.cell(row=rwCount1, column=3).value)
    userStory_PointR2 = int(worksheet2.cell(row=rwCount2, column=3).value)
    if worksheet2.cell(row=rwCount2, column=5).value is None:
        userStory_Pointadd = 0
    else:
        usp = worksheet2.cell(row=rwCount2, column=5).value
        userStory_Pointadd = int(usp)

    userStory_Pointadd = userStory_PointR1 + userStory_PointR2 + userStory_Pointadd
    cell = worksheet2.cell(row=rwCount2, column=5)  # writing the new value on R2 sheet: repeat of R2 is getting calculated
    writeto_excel(cell, userStory_Pointadd)


def valuesCarriedToR1_WeightedSim(rwCount1, rwCount2, similarityVal):  # configuration=4
    userStory_PointR2 = int(worksheet2.cell(row=rwCount2, column=3).value)
    if worksheet1.cell(row=rwCount1, column=3).value is None:
        userStory_PointR1 = 0
    else:
        usp = worksheet1.cell(row=rwCount1, column=3).value
        userStory_PointR1 = float(usp)

    userStory_PointR1 = userStory_PointR1 + (similarityVal * userStory_PointR2)
    cell = worksheet1.cell(row=rwCount1, column=3)
    writeto_excel(cell, userStory_PointR1)


def texualcomparison(configuration):

    maxr1 = worksheet1.max_row
    maxr2 = worksheet2.max_row
    print(f"r1: {maxr1} r2: {maxr2}")
    mat = [[0 for i in range(maxr1 - 1)] for j in range(maxr2 - 1)]
    rwCount1 = 1
    rwCount2 = 1
    similarity_threshold = 0.85
    mappingValue = ""
    for row2 in worksheet2.iter_rows(min_row=2, max_row=maxr2, min_col=2, max_col=2):
        for cell2 in row2:
            r2 = cell2.value
            rwCount2 += 1
            if rwCount2 <= 2:
                pass
            else:
                rwCount1 = 1
            mappingValue = ""
            for row1 in worksheet1.iter_rows(min_row=2, max_row=maxr1, min_col=2, max_col=2):
                for cell1 in row1:
                    r1 = cell1.value
                    rwCount1 += 1
                    print(f"R2: {rwCount2} and R1: {rwCount1}")
                    similarity_score = textsimilarity(r1, r2)
                    mat[rwCount2 - 2][rwCount1 - 2] = similarity_score
                    if similarity_score > similarity_threshold:
                        # print(f"score: {similarity_score} ")      ********
                        # f"r1:{r1} and r2:{r2}")
                        if (configuration == 1):
                            valuesCarriedToR1(rwCount1, rwCount2)
                        elif (configuration == 2):
                            valuesCarriedToR2(rwCount1, rwCount2)
                        elif (configuration == 3):
                            valuesR1andR2Added(rwCount1, rwCount2)
                        elif (configuration == 4):
                            valuesCarriedToR1_WeightedSim(rwCount1, rwCount2, similarity_score)

                        mapVal2 = worksheet2.cell(row=rwCount2, column=1).value
                        cell = worksheetMap.cell(row=rwCount2, column=1)
                        writeto_excel(cell, mapVal2)

                        mapVal1 = worksheet1.cell(row=rwCount1, column=1).value
                        if len(mappingValue) == 0:
                            mappingValue += str(mapVal1)
                        else:
                            mappingValue = mappingValue + ',' + str(mapVal1)
                        cell = worksheetMap.cell(row=rwCount2, column=2)
                        writeto_excel(cell, mappingValue)

                        # print(f"Updated userStory_PointR1: {userStory_PointR1}")

                    else:
                        # print(f"Similarity value below the threshold {similarity_threshold}")     **********
                        mapVal2 = worksheet2.cell(row=rwCount2, column=1).value
                        cell = worksheetMap.cell(row=rwCount2, column=1)
                        writeto_excel(cell, mapVal2)

                        # cell = worksheetMap.cell(row=rwCount2, column=2)
                        # writeto_excel(cell, "No Match")

    print(mat)
    writematrix(workbook, mat)
    buildmatrixsheet(workbook)
    writeuspointsto_matsheet(workbook)



# textn1="As a developer, I'd like to move k8s SPI to it's own repo."
# textn2="As a developer, I'd like to upgrade Spring XD's ambari plugin to 1.3 release."
#
# print(textsimilarity(textn1,textn2))
