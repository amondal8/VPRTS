import mysql.connector
import openpyxl as op
import configparser
import json




file = "ini1"
config = configparser.ConfigParser()
config.read('config1.ini')
dbconnect = config['dbconnection_dataset']
data = config['data']

mydb = mysql.connector.connect(
  host=dbconnect["host"],
  user=dbconnect["user"],
  password=dbconnect["password"],
  database=dbconnect["database"]
)



filepath = "/Thesis/Files/Extra/Database Creation/Mapping.xlsx"
workbook = op.load_workbook(filepath)

filepath_saveresults = "C:/Users/amondal8/PycharmProjects/pythonProject3/Thesis/Files/Definitional data/My_results.xlsx"
workbook_res = op.load_workbook(filepath_saveresults)
saveresults_sheetname = "new"


  # "Simval_Distds_id"

  # "RQ1_Run4"


def getds_id():
  print(dbconnect["database"])
  mycursor = mydb.cursor()
  mycursor.execute("Select ds_id from dataset ORDER BY timestamp DESC LIMIT 1")
  result = mycursor.fetchone()
  if result is None:
    return result
  else:
    return result[0]


def stringcreation_columns(my_list):
  my_str = ""
  for ind, i in enumerate(my_list):
    if ind == 0 and len(my_list) != 1:
      my_str = my_str + "(" + str(i)
    elif ind ==0 and len(my_list) == 1:
      my_str = my_str + "(" + str(i) + ")"
    elif ind == len(my_list) - 1:
      my_str = my_str + "," + str(i) + ")"
    else:
      my_str = my_str + "," + str(i)
  return my_str


def stringcreation_values(my_list, datatype_list):
  my_str = ""
  for ind, i in enumerate(my_list):
    if ind == 0 and len(my_list) != 1:
      if datatype_list[ind] == "int":
        my_str = my_str + "(" + str(i)
      else:
        my_str = my_str + "(\"" + str(i) + "\""
    elif ind == 0 and len(my_list) == 1:
      if datatype_list[ind] == "int":
        my_str = my_str + "(" + str(i) + ")"
      else:
        my_str = my_str + "(\"" + str(i) + "\")"
    elif ind == len(my_list) - 1:
      if datatype_list[ind] == "int":
        my_str = my_str + "," + str(i) + ")"
      else:
        my_str = my_str + ",\"" + str(i) + "\")"
    else:
      if datatype_list[ind] == "int":
        my_str = my_str + "," + str(i)
      else:
        my_str = my_str + ",\"" + str(i) + "\""
  return my_str


def running_insertquery(query):
  mydb = mysql.connector.connect(
    host=dbconnect["host"],
    user=dbconnect["user"],
    password=dbconnect["password"],
    database=dbconnect["database"]
  )
  mycursor = mydb.cursor()
  if mydb.is_closed():
    mydb.start_transaction()
  # mycursor = cc.connection()
  mycursor.execute(query)
  # print(f"Query {query} executed successfully")
  mydb.commit()
  # cc.closeconnection()
  mycursor.close()
  mydb.close()


def createdictfromexcel(filepath, worksheetName, keycol, valuecol):
  workbook = op.load_workbook(filepath)
  worksheet = workbook[worksheetName]
  maxr = worksheet.max_row
  keyval_dict = {}

  for row_ind, row in enumerate(worksheet.iter_rows(min_row=1, max_row=maxr, min_col=keycol, max_col=keycol),
                                start=1):
    for cell in row:
      valcell = worksheet.cell(row=row_ind, column=valuecol)
      val = valcell.value
      # if value is None:
      #     value=None
      keyval_dict[cell.value] = val
  # print(keyval_dict)
  return keyval_dict


def fetch_ones_in_row(adjacency_matrix, row_index):
  row = adjacency_matrix[row_index]
  ones_indices = [index for index, value in enumerate(row) if value == 1]
  return ones_indices


def running_searchqury(query):
  mycursor = mydb.cursor()
  mycursor.execute(query)
  result = mycursor.fetchall()
  # print(result)
  return result

def create_fetchquery_usvalue(colname, tc_id, ds_id):
  query = f"""Select sum({colname}) from userstory_datasettable where us_id in(SELECT DISTINCT us_id from us_tc_map where tc_id in ("{tc_id}") and ds_id = {ds_id}) and ds_id = {ds_id}"""
  # print(query)
  return query


def create_fetchquery_tcexectime(colname, tc_id, ds_id):
  query = f"""SELECT {colname} from tc_datasettable where tc_id = "{tc_id}" and ds_id = "{ds_id}"; """
  print(query)
  return query


def creating_prioritydict_tclist(tc_list, ds_id):
  tc_dict = {}
  for ind, i in enumerate(tc_list):
    query_usp = create_fetchquery_usvalue("us_points", i, ds_id)
    query_usbv = create_fetchquery_usvalue("us_businessvalue", i, ds_id)
    # print(query)
    result1 = running_searchqury(query_usp)
    result2 = running_searchqury(query_usbv)
    tc_dict[i] = int(result1[0][0]) + int(result2[0][0])
  sorted_tcdict = dict(sorted(tc_dict.items(), key=lambda x: x[1], reverse=True))
  print(f"Printing the selected tc set with usp: {tc_dict}")
  print(f"Printing the selected set in decreasing order of usp: {sorted_tcdict}")
  return sorted_tcdict


def ini_to_json(ini_filepath):
  # Read the .ini file
  config = configparser.ConfigParser()
  config.read(ini_filepath)

  # Convert the .ini file to a dictionary
  ini_dict = {}
  for section in config.sections():
    ini_dict[section] = dict(config[section])

  # Convert the dictionary to a JSON string
  json_string = json.dumps(ini_dict, indent=1)

  return json_string


def insert_json_data(json_data, ds_id):
  sql_insert_query = f"UPDATE dataset SET ini_file = ('{json_data}') where ds_id = '{ds_id}'"
  # print(sql_insert_query)
  running_insertquery(sql_insert_query)


def json_to_ini(json_data):
  data_dict = json.loads(json_data)
  # Add each key-value pair to the 'DEFAULT' section of the .ini file
  for key, value in data_dict.items():
    config[key] = value
  with open('copiedconfig.ini', 'w') as configfile:
    config.write(configfile)


def read_result(ds_id):
  query = f"Select results from dataset where ds_id = '{ds_id}'"
  res = running_searchqury(query)
  print(f"Result: \n{res[0][0]}")


def read_json(ds_id):
  query = f"Select ini_file from dataset where ds_id = '{ds_id}'"
  res = running_searchqury(query)
  print(f"config file: \n{res[0][0]}")


def getvalue_fromcol(matrix, colind):
  column_value_list = []

  for row in matrix:
    val = row[colind]
    column_value_list.append(val)
  return column_value_list

def normalized_list(raw_list):
  normalizer = 1/sum(raw_list)
  norm_list = [float(i) * normalizer for i in raw_list]
  return norm_list

# def normalized_list(raw_list):
#   min_value = min(raw_list)
#   max_value = max(raw_list)
#   scaled = [(x - min_value) / (max_value - min_value) for x in raw_list]
#   return scaled

def simval_config(mat, us1_list):
  simval_dict = {}
  simval_list = []
  for indj, j in enumerate(us1_list):
      val_list = getvalue_fromcol(mat, indj)
      sum_val = sum(val_list)
      simval_list.append(sum_val)
  norm_simval = normalized_list(simval_list)
  # print(f"sum of normed: {sum(norm_simval)}")
  for indj, j in enumerate(us1_list):
    simval_dict[j] = norm_simval[indj]
  return simval_dict


def importanceval_fortc(ds_id, mat, us2_list, us1_list, runconfig):
  print(us1_list)
  print(us2_list)
  bvvalr1list_raw = []
  bvvalr2list_raw = []
  cmvallist_raw = []
  impval_dict = {}
  for i in us1_list:
    query = f"Select us_businessvalue from userstory_datasettable where us_id = '{i}' and ds_id = '{ds_id}'"
    res = running_searchqury(query)
    bvvalr1list_raw.append(res[0][0])
  bvvalr1_list = normalized_list(bvvalr1list_raw)
  for i in us2_list:
    query1 = f"Select us_businessvalue from userstory_datasettable where us_id = '{i}' and ds_id = '{ds_id}'"
    query2 = f"select sum(affected_value) from us_cm_map where us_id in ('{i}') and ds_id = '{ds_id}' group by us_id;"
    res1 = running_searchqury(query1)
    res2 = running_searchqury(query2)
    bvvalr2list_raw.append(res1[0][0])
    cmvallist_raw.append(res2[0][0])
  cmval_list = normalized_list(cmvallist_raw)
  bvvalr2_list = normalized_list(bvvalr2list_raw)
  print(f"normalized business value: {bvvalr2_list}")
  # print(f"cmval_list:{cmval_list}, sum of the list:{sum(cmval_list)}")
  # print(f"bvvalr2_list:{bvvalr2_list}, sum of the list:{sum(bvvalr2_list)}")
  for indj, j in enumerate(us1_list):
      val_list = getvalue_fromcol(mat, indj)
      # print(f"Sim val_list: {val_list}")
      importance_val=0
      for indk, k in enumerate(val_list):
        if runconfig == "1.1":          #Considering bv of current user stories
          importance_val += k*bvvalr2_list[indk]
        elif runconfig == "1.2":        #Considering bv of previous user stories
          importance_val += k*bvvalr1_list[indj]
        elif runconfig == "1.3":        #Combining bv of current and previous user stories
          importance_val += k * bvvalr2_list[indk] * bvvalr1_list[indj]
        elif runconfig == "0":          #Calculating only sim val: This is for comparing with our result
          simval_dict = simval_config(mat, us1_list)
          sorted_simvaldict = {k: v for k, v in sorted(simval_dict.items(), key=lambda item: item[1], reverse=True)}
          return sorted_simvaldict
        elif runconfig == "2":          # This is combining code modification with 1.1
          importance_val += k*cmval_list[indk]*bvvalr2_list[indk]
      impval_dict[j] = importance_val
  # print(impval_dict)
  return impval_dict


def writeconfigs_to_excel(colindex, col, value):

  worksheet_res = workbook_res[saveresults_sheetname]
  column = worksheet_res[col]
  # print(f"len for col{col}: {len(column)}")

  if colindex == "next":
    next_row = len(column) + 1
  else:
    next_row = len(column)
  cell = worksheet_res[f'{col}{next_row}']
  cell.value = value
  workbook_res.save(filepath_saveresults)


def readandwrite_toexcel(readval, col, writeval):
  worksheet_res = workbook_res[saveresults_sheetname]
  row = 0
  readcol = "A"
  column = worksheet_res[readcol]
  for i in range(len(column)):
    cell = worksheet_res[f'{readcol}{i+1}']
    if str(cell.value) == str(readval):
      row = i+1
      break
  print(f"Row: {row}, col: {col}")
  cell1 = worksheet_res[f'{col}{row}']
  cell1.value = writeval
  workbook_res.save(filepath_saveresults)


def commitconnection():
  mydb.commit()

def creatingset_fromstring(s):
  new_set = set()
  s = s.replace("{", "").replace("}", "").replace("'","")
  s_list = s.split(",")
  for i in s_list:
    new_set.add(i.strip())
  return new_set


def creatingstring_fromlist(ls):
  s = str(ls).replace("[", "").replace("]", "")
  return s

def knapsack_01(values, weights, W):
  # Number of items
  n = len(values)

  # Create a DP table to store results of subproblems
  dp = [[0 for x in range(W + 1)] for y in range(n + 1)]

  # Fill dp[][] in bottom-up manner
  for i in range(1, n + 1):
    for w in range(1, W + 1):
      if weights[i - 1] <= w:
        # Max of two cases:
        # (1) nth item included
        # (2) not included
        dp[i][w] = max(dp[i - 1][w], dp[i - 1][w - weights[i - 1]] + values[i - 1])
      else:
        # If the weight of the nth item is more than Knapsack of capacity w,
        # then this item cannot be included in the optimal solution
        dp[i][w] = dp[i - 1][w]

  # Return the maximum value that can be put in a knapsack of capacity W
  return dp[n][W]




# read_result(1)
# read_json(1)

# writeconfigs_to_excel("next", 'A', "Value1")
# writeconfigs_to_excel("same", 'B', "Value2")


# json_data = ini_to_json("config1.ini")
# print(json_data)
# # insert_json_data(json_data, 1)
# json_to_ini(json_data)
# json_data = ini_to_json("copiedconfig.ini")
# print(json_data)

# creating_prioritydict_tclist(["TC1", "TC2"], "1")

# us_lst = [1,2,3,4]
# create_tcsearchquery(us_lst, "tc_id", "us_tc_map", "us_id")